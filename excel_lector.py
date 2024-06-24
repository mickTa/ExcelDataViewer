import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from PIL import Image, ImageTk

class ExcelDataViewer(tk.Frame):
    def __init__(self, parent, excel_path, image_path):
        super().__init__(parent)
        self.parent = parent
        self.excel_path = excel_path
        self.image_path = image_path
        self.workbook = None
        self.all_rows = []

        self.init_ui()
        self.load_excel_file()

    def init_ui(self):
        image = Image.open(self.image_path)
        photo = ImageTk.PhotoImage(image)

        file_frame = ttk.LabelFrame(self)
        file_frame.pack(padx=10, pady=5)

        image_label = ttk.Label(file_frame, image=photo)
        image_label.image = photo  
        image_label.pack()

        search_frame = ttk.LabelFrame(self, text="Recherche")
        search_frame.pack(fill=tk.X, padx=10, pady=5)

        search_label = ttk.Label(search_frame, text="Recherche : ")
        search_label.pack(side=tk.LEFT, padx=5, pady=5)

        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)

        search_button = ttk.Button(search_frame, text="Rechercher", command=self.search_data)
        search_button.pack(side=tk.LEFT, padx=5, pady=5)

        data_frame = ttk.LabelFrame(self, text="Données")
        data_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.data_display = ttk.Treeview(data_frame, show='headings')
        self.data_display.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Add a scrollbar
        scrollbar = ttk.Scrollbar(self.data_display, orient=tk.VERTICAL, command=self.data_display.yview)
        self.data_display.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def load_excel_file(self):
        if self.excel_path:
            self.workbook = load_workbook(self.excel_path)
            sheet_names = self.workbook.sheetnames
            self.display_sheet_data(sheet_names[0])  

    def display_sheet_data(self, sheet_name):
        if sheet_name and self.workbook:
            sheet = self.workbook[sheet_name]
            for row in self.data_display.get_children():
                self.data_display.delete(row)

            columns = [cell.value for cell in sheet[1]]
            self.data_display['columns'] = columns
            for col in columns:
                self.data_display.heading(col, text=col)
                self.data_display.column(col, width=100, anchor=tk.CENTER)

            self.all_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.data_display.insert('', tk.END, values=row)
                self.all_rows.append(row)

    def search_data(self):
        keyword = self.search_entry.get().lower()
        filtered_rows = [row for row in self.all_rows if any(keyword in str(cell).lower() for cell in row)]

        for row in self.data_display.get_children():
            self.data_display.delete(row)

        for row in filtered_rows:
            self.data_display.insert('', tk.END, values=row)
